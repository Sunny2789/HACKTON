from tkinter import *
from tkinter import messagebox
import sqlite3
import os
import time
import openpyxl as op

root = Tk()
root.title("Health Care Management System")
root.iconbitmap('Office.ico')
root.geometry("5000x800")

presentDate = time.asctime(time.localtime(time.time()))

def Add_Patient():
    headName.set("Add Patient")
    firstLabel.set("Patient's Name")
    secondLabel.set("Patient's Age")
    thirdLabel.set("Patient's IP.NO.")
    fourthLabel.set("Patient's Address")
    fivthLabel.set("Patient's Phone")
    sixthLabel.set("Date Of Admission")

    firstEntry.set("")
    secondEntry.set("")
    thirdEntry.set("")
    fourthEntry.set("")
    fivthEntry.set("")
    sixthEntry.set(presentDate)

def Remove_Patient():
    headName.set("Remove Patient")
    firstLabel.set("Patient's Name")
    secondLabel.set("Patient's IP.No")
    thirdLabel.set("")
    fourthLabel.set("")
    fivthLabel.set("")
    sixthLabel.set("")

    thirdEntry.set("Not Required")
    fourthEntry.set("Not Required")
    fivthEntry.set("Not Required")
    sixthEntry.set("Not Required")

def Appointment_Generation():
    headName.set("Appointment Generation")
    firstLabel.set("Patient's Name")
    secondLabel.set("Patient's Age")
    thirdLabel.set("Patient's IP.No.")
    fourthLabel.set("Consulting Doctor")
    fivthLabel.set("Appointment Date")
    sixthLabel.set("Appointment Details")

    firstEntry.set("")
    secondEntry.set("")
    thirdEntry.set("")
    fourthEntry.set("")
    fivthEntry.set("")
    sixthEntry.set("")

def submit():
    path = os.getcwd()
    cwd = path.replace("\\", "/")

    f_entry = firstEntry.get()
    se_entry = secondEntry.get()
    t_entry = thirdEntry.get()
    fo_entry = fourthEntry.get()
    fi_entry = fivthEntry.get()
    s_entry = sixthEntry.get()

    if headName.get()=="Add Patient":
        if f_entry.isalpha()==False:
            messagebox.showerror("Patient Name Error", "Patient name should not contain any space or number.")
        if se_entry.isnumeric()==False:
                messagebox.showerror("Patient Age Error", "Patient age should not contain any alphabets.")
        if fi_entry.isnumeric()==False or len(fi_entry)<10 or len(fi_entry)>10:
                messagebox.showerror("Patient Phone Error", "Patient phone should only contain numbers and it should not be less than or more than 10/")

        elif f_entry.isalpha()==True and se_entry.isnumeric()==True and fi_entry.isalpha()==False and len(fi_entry)==10:
            statusVar.set("Getting Patient Information")
            status.update()
            time.sleep(2)

            IpNo =t_entry

            try:
                folder = "Patient"+IpNo
                os.makedirs(folder)

                statusVar.set("Making Hospital Files For Patient")
                status.update()
                time.sleep(2)

                dataBase = sqlite3.connect('patient_database.db')
                cursor = dataBase.cursor()



                cursor.execute("INSERT INTO patients VALUES (:name, :age, :phone, :address, :admission)",
                                {
                                'name': f_entry,
                                'age': se_entry,
                                'phone': fi_entry,
                                'address': fo_entry,
                                'admission': s_entry
                                })
                cursor.execute("SELECT *, oid FROM patients")
                records = cursor.fetchall()
                print(records)

                dataBase.commit()
                dataBase.close()

                with open(folder+"/patient_info.txt", 'w') as f:
                    f.write("Patient Name: ")
                    f.write(f_entry)
                    f.write("\nPatient Age:")
                    f.write(se_entry)
                    f.write("\nPatient IP.No: ")
                    f.write(t_entry)
                    f.write("\nPatient Address: ")
                    f.write(fo_entry)
                    f.write("\nPatient Phone: ")
                    f.write(fi_entry)
                    f.write("\nPatient Date Of Admission: ")
                    f.write(s_entry)

                workbook = op.load_workbook('patients_file.xlsx')
                sheet1 = workbook['Sheet1']

                max_r = sheet1.max_row

                data1 = sheet1.cell(row=1, column=1).value
                data2 = sheet1.cell(row=1, column=2).value
                data3 = sheet1.cell(row=1, column=3).value

                data4 = sheet1.cell(row=1, column=4).value
                data5 = sheet1.cell(row=1, column=6).value

                data6 = sheet1.cell(row=1, column=7).value

                if ((data1 != 'IP.No' and data2 != 'Patient Name' and data3 != 'Age' and data4 != 'Phone')
                    and (data5 != 'Address' and data6 != 'Admission Date')):
                    sheet1.cell(row=1, column=1).value = "IP.No"
                    sheet1.cell(row=1, column=2).value = "Patient Name"
                    sheet1.cell(row=1, column=3).value = "Age"
                    sheet1.cell(row=1, column=4).value = "Phone"
                    sheet1.cell(row=1, column=5).value = "Address"
                    sheet1.cell(row=1, column=6).value = "Admission Date"
                else:
                    pass

                sheet1.cell(row=max_r+1, column=1).value=t_entry
                sheet1.cell(row=max_r+1, column=2).value=f_entry
                sheet1.cell(row=max_r+1, column=3).value=se_entry
                sheet1.cell(row=max_r+1, column=4).value=fi_entry
                sheet1.cell(row=max_r+1, column=5).value=fo_entry
                sheet1.cell(row=max_r+1, column=6).value=s_entry

                workbook.save('patients_file.xlsx')

                statusVar.set("Patient Creation Successfull")
                status.update()
                firstEntry.set("")
                secondEntry.set("")
                thirdEntry.set("")
                fourthEntry.set("")
                fivthEntry.set("")
                sixthEntry.set(presentDate)

                messagebox.showinfo("Patient Creation SuccessFull", "The New Patient Files Have Been Created.")

            except Exception as e:
                messagebox.showerror("Patient Existing Error", "The patient cannot be created because the patient with the given IP.No already exists.")




# Submenus
mainmenu = Menu(root)

m1 = Menu(mainmenu, tearoff = 0)
m1.add_command(label = "Open A File")
m1.add_command(label = "Get Information's")
m1.add_separator()
m1.add_command(label = "Take Out Print")
m1.add_command(label = "Exit", command = quit)
root.config(menu = mainmenu)
mainmenu.add_cascade(label = "File", menu = m1)

m2 = Menu(mainmenu, tearoff = 0)
m2.add_command(label = "New Form")

root.config(menu = mainmenu)
mainmenu.add_cascade(label = "Edit", menu = m2)

# ----------------------------------------------------------------------------------

sideFrame = Frame(root, bg="powder blue", padx=20, borderwidth=2, relief=RIDGE)
sideFrame.pack(side=LEFT, fill=Y)

headFrame = Frame(root, bg="powder blue", borderwidth=2, relief=SUNKEN)
headFrame.pack(side=TOP, fill=X)

#--------------------------------------------------------------------------------------

l1 = Label(headFrame, text="Health Management System", font="Helvetica 20 bold")
l1.pack()

l2 = Label(sideFrame, text="Process", font="Helvetica 16 bold")
l2.pack(fill=X)

#------------------------------------------------------------------------------------

patientFrame = Frame(sideFrame, borderwidth=5, relief=RAISED, padx=30)
patientFrame.pack(pady=10)

receiptFrame = Frame(sideFrame, borderwidth=5, relief=RAISED)
receiptFrame.pack(pady=10)

showFrame = Frame(sideFrame, borderwidth=5, relief=RAISED)
showFrame.pack(pady=10)

#--------------------------------------------------------------------------------------
# Buttons And Labels of Side Frame
l3 = Label(patientFrame, text="Patients Process", bg="powder blue", font="Arial 18 bold", padx=50)
l3.pack(pady=20)
ppb1 = Button(patientFrame, text="Add Patient", bg="wheat", font="ArialRounded 14 bold", padx=90, command=Add_Patient)
ppb1.pack(pady=10, fill=X)
ppb2 = Button(patientFrame, text="Remove Patient", bg="wheat", font="ArialRounded 14 bold", padx=90, command=Remove_Patient)
ppb2.pack(fill=X)

l4 = Label(receiptFrame, text="Billing Process", bg="powder blue", font="Arial 18 bold", padx=50)
l4.pack(pady=20)
ppb4 = Button(receiptFrame, text="Generate Appointment", bg="wheat", font="ArialRounded 14 bold", padx=90, command=Appointment_Generation)
ppb4.pack(pady=5, fill=X)
ppb5 = Button(receiptFrame, text="Generate Prescription", bg="wheat", font="ArialRounded 14 bold", padx=90)
ppb5.pack(pady=5, fill=X)

l5 = Label(showFrame, text="Details Process", bg="powder blue", font="Arial 18 bold", padx=50)
l5.pack(pady=20)
ppb6 = Button(showFrame, text="Get Patient Info", bg="wheat", font="ArialRounded 14 bold", padx=90)
ppb6.pack(pady=5, fill=X)
ppb7 = Button(showFrame, text="Get PatientAppointment", bg="wheat", font="ArialRounded 14 bold", padx=90)
ppb7.pack(pady=5, fill=X)
ppb8 = Button(showFrame, text="Get PatientPrescription", bg="wheat", font="ArialRounded 14 bold", padx=90)
ppb8.pack(pady=5, fill=X)

# -------------------------------------------------------------------------------------
headName = StringVar()

firstEntry = StringVar()
secondEntry = StringVar()
thirdEntry = StringVar()
fourthEntry = StringVar()
fivthEntry = StringVar()
sixthEntry = StringVar()

firstLabel = StringVar()
secondLabel = StringVar()
thirdLabel = StringVar()
fourthLabel = StringVar()
fivthLabel = StringVar()
sixthLabel = StringVar()

bodyFrame = Frame(root, borderwidth=5, relief=RAISED)
bodyFrame.pack(fill=X, pady=20)

bheadFrame = Frame(bodyFrame, bg="coral")
bheadFrame.pack(fill=X)

formFrame = Frame(bodyFrame, bg="grey", padx=50)
formFrame.pack()

bFL1 = Label(bheadFrame, textvariable=headName, font="Helvetica 20 bold")
bFL1.pack()

Label(formFrame, textvariable=firstLabel, font="ComicSan 18 bold", borderwidth=3, relief=RAISED).grid(row=0, column=0, pady=25, sticky=W)
Label(formFrame, textvariable=secondLabel, font="ComicSan 18 bold", borderwidth=3, relief=RAISED).grid(row=1, column=0, pady=25, sticky=W)
Label(formFrame, textvariable=thirdLabel,  font="ComicSan 18 bold", borderwidth=3, relief=RAISED).grid(row=2, column=0, pady=25, sticky=W)
Label(formFrame, textvariable=fourthLabel, font="ComicSan 18 bold", borderwidth=3, relief=RAISED).grid(row=3, column=0, pady=25, sticky=W)
Label(formFrame, textvariable=fivthLabel, font="ComicSan 18 bold", borderwidth=3, relief=RAISED).grid(row=4, column=0, pady=25, sticky=W)
Label(formFrame, textvariable=sixthLabel, font="ComicSan 18 bold", borderwidth=3, relief=RAISED).grid(row=5, column=0, pady=25, sticky=W)

first_Entry = Entry(formFrame, textvariable=firstEntry, width=30, borderwidth=8, font="ComicSan 12 bold", relief=RAISED).grid(row=0, column=1, padx=50)
second_Entry = Entry(formFrame, textvariable=secondEntry, width=30, borderwidth=8, font="ComicSan 12 bold", relief=RAISED).grid(row=1, column=1, padx=50)
third_Entry = Entry(formFrame, textvariable=thirdEntry, width=30, borderwidth=8, font="ComicSan 12 bold", relief=RAISED).grid(row=2, column=1, padx=50)
fourth_Entry= Entry(formFrame, textvariable=fourthEntry, width=30, borderwidth=8, font="ComicSan 12 bold", relief=RAISED).grid(row=3, column=1, padx=50)
fivth_Entry= Entry(formFrame, textvariable=fivthEntry, width=30, borderwidth=8, font="ComicSan 12 bold", relief=RAISED).grid(row=4, column=1, padx=50)
sixth_Entry= Entry(formFrame, textvariable=sixthEntry, width=30, borderwidth=8, font="ComicSan 12 bold", relief=RAISED).grid(row=5, column=1, padx=50)

submitButton = Button(formFrame, text="Submit", font="ArialRounded 15 bold", padx=80, bg="powder blue", borderwidth=10, relief=RAISED, command=submit)
submitButton.grid(row=6, column=0, columnspan=1, pady=20)

cancelButton = Button(formFrame, text="Cancel", font="ArialRounded 15 bold", padx=80, bg="powder blue", borderwidth=10, relief=RAISED)
cancelButton.grid(row=6, column=1, columnspan=1)

# --------------------------------------------------------------------------------------
statusFrame = Frame(root, bd=3, relief=SUNKEN)
statusFrame.pack(fill=BOTH, side=BOTTOM)

statusLabel = Label(statusFrame, text="Status", borderwidth=5, relief=GROOVE, font="TimesNewRoman 12 bold")
statusLabel.grid(row=0, column=0, sticky=W)

statusVar = StringVar()
statusVar.set("Ready")

status = Label(statusFrame, textvariable=statusVar, font="TimesNewRoman 12")
status.grid(row=1, column=0)
root.mainloop()
